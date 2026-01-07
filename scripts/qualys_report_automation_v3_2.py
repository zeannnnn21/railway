#!/usr/bin/env python3
"""
Qualys Vulnerability Report Automation Script v3.2 - Enhanced Version
Automatically populates detailed reports from Qualys CSV scan results
Supports single file processing and batch processing

Key improvements in v3.2:
- Added monitoring/logging Excel file generation
- Automatically extracts scan metadata for tracking
- Creates copy-paste ready monitoring log

Key improvements in v3.1:
- Summary sheet and chart now preserved from template (no chart recreation)
- Chart automatically updates from Severity Tabulation data
- Simplified workflow - just update data, chart refreshes automatically

Key improvements in v3.0:
- Fixed Business Unit field update in Details sheet
- Added AutoFilter to Vulnerability Details headers
- Improved text formatting and readability
- Better row height management
- Enhanced error handling and validation
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
import re
import sys
import os
from datetime import datetime
from pathlib import Path

class QualysReportAutomation:
    """Main class for automating Qualys report generation"""
    
    SEVERITY_MAP = {
        '1': 'Low',
        '2': 'Medium',
        '3': 'High',
        '4': 'Critical',
        '5': 'Critical'
    }
    
    SEVERITY_ORDER = {
        'Critical': 1,
        'High': 2,
        'Medium': 3,
        'Low': 4,
        'Unknown': 5
    }
    
    CVSS_MAP = {
        '1': '3',
        '2': '5',
        '3': '7',
        '4': '9',
        '5': '10'
    }
    
    def __init__(self):
        self.stats = {
            'total_files': 0,
            'successful': 0,
            'failed': 0,
            'total_vulnerabilities': 0
        }
    
    @staticmethod
    def parse_severity(severity_str):
        """Map Qualys severity to standard levels"""
        # Handle float values (e.g., 2.0, 3.0) by converting to int first
        try:
            if pd.notna(severity_str):
                severity_int = str(int(float(severity_str)))
                return QualysReportAutomation.SEVERITY_MAP.get(severity_int, 'Unknown')
        except (ValueError, TypeError):
            pass
        return QualysReportAutomation.SEVERITY_MAP.get(str(severity_str).strip(), 'Unknown')
    
    @staticmethod
    def get_cvss_score(severity_str):
        """Estimate CVSS score from severity level"""
        return QualysReportAutomation.CVSS_MAP.get(str(severity_str).strip(), '')
    
    @staticmethod
    def format_port_protocol(port, protocol):
        """Format port and protocol as port/protocol"""
        if pd.notna(port) and pd.notna(protocol):
            # Handle float ports from CSV
            port_val = str(port).replace('.0', '') if isinstance(port, float) else str(port)
            return f"{port_val}/{protocol}"
        elif pd.notna(port):
            port_val = str(port).replace('.0', '') if isinstance(port, float) else str(port)
            return port_val
        return ''
    
    @staticmethod
    def generate_cve_urls(cve_ids):
        """Generate NIST NVD URLs for CVE IDs"""
        if pd.isna(cve_ids) or not cve_ids:
            return ''
        
        cve_list = [cve.strip() for cve in str(cve_ids).split(',')]
        urls = []
        for cve in cve_list:
            if cve and 'CVE-' in cve:
                urls.append(f"https://nvd.nist.gov/vuln/detail/{cve}")
        
        return '\n'.join(urls) if urls else ''
    
    @staticmethod
    def count_exploits(exploitability, bugtraq_id):
        """Count potential exploits based on available data"""
        count = 0
        if pd.notna(exploitability) and str(exploitability).strip():
            count += 1
        if pd.notna(bugtraq_id) and str(bugtraq_id).strip():
            count += 1
        return str(count) if count > 0 else ''
    
    @staticmethod
    def clean_text(text):
        """Clean and format text fields with improved readability"""
        if pd.isna(text):
            return ''
        text = str(text).strip()
        # Remove excessive whitespace
        text = re.sub(r'\s+', ' ', text)
        # Add line breaks for better readability after sentences
        text = re.sub(r'\.\s+([A-Z])', r'.\n\n\1', text)
        return text
    
    @staticmethod
    def extract_scan_info(csv_file):
        """Extract scan information from Qualys CSV header"""
        info = {
            'scan_title': '',
            'asset_groups': '',
            'ips': '',
            'scan_date': ''
        }
        
        try:
            with open(csv_file, 'r', encoding='utf-8') as f:
                lines = [f.readline() for _ in range(7)]
                
            # Parse scan info from line 6 (index 5)
            if len(lines) > 5:
                parts = lines[5].strip('"').split('","')
                if len(parts) >= 9:
                    info['scan_date'] = parts[0]
                    info['scan_title'] = parts[8]
                    info['asset_groups'] = parts[9] if len(parts) > 9 else ''
                    info['ips'] = parts[10] if len(parts) > 10 else ''
        except Exception as e:
            print(f"⚠ Warning: Could not extract scan info: {e}")
        
        return info
    
    def process_qualys_csv(self, csv_file):
        """Read and process Qualys CSV file"""
        # Read CSV, skip header rows
        df = pd.read_csv(csv_file, skiprows=7, encoding='utf-8')
        
        # Filter out rows where IP is empty
        df = df[df['IP'].notna()]
        
        # Filter to only include "Vuln" type (exclude "Practice", "Ig", "Potential")
        df = df[df['Type'] == 'Vuln']
        
        return df
    
    def update_details_sheet(self, wb, business_unit):
        """Update the Details sheet with business unit information"""
        if 'Details' not in wb.sheetnames:
            print("⚠ Warning: 'Details' sheet not found in template")
            return
        
        ws = wb['Details']
        
        try:
            # Update Row 9, Column 3 with Business Unit name
            # Row 9, Col 2 should have "Business Unit / Group:"
            if ws.cell(row=9, column=2).value and 'Business Unit' in str(ws.cell(row=9, column=2).value):
                ws.cell(row=9, column=3, value=business_unit)
                print(f"✓ Details sheet: Business Unit set to '{business_unit}' (Row 9, Col 3)")
            
            # Update Row 15, Column 2 - Replace <Business Unit / Group> placeholder
            cell_15_2 = ws.cell(row=15, column=2)
            if cell_15_2.value and '<Business Unit / Group>' in str(cell_15_2.value):
                cell_15_2.value = str(cell_15_2.value).replace('<Business Unit / Group>', business_unit)
                print(f"✓ Details sheet: Replaced placeholder in Row 15")
            
            # Update Date of Submission (Row 10, Column 3)
            if ws.cell(row=10, column=2).value and 'Date of Submission' in str(ws.cell(row=10, column=2).value):
                ws.cell(row=10, column=3, value=datetime.now().strftime("%B %d, %Y"))
                print(f"✓ Details sheet: Date updated to {datetime.now().strftime('%B %d, %Y')}")
                
        except Exception as e:
            print(f"⚠ Warning: Error updating Details sheet: {e}")
    
    def fill_vulnerability_details(self, template_file, csv_file, output_file, business_unit=None):
        """Fill the template with vulnerability data from CSV"""
        
        try:
            print(f"\n{'='*60}")
            print(f"Processing: {Path(csv_file).name}")
            print(f"{'='*60}")
            
            # Extract scan information
            scan_info = self.extract_scan_info(csv_file)
            
            # Read vulnerability data
            vuln_df = self.process_qualys_csv(csv_file)
            print(f"✓ Found {len(vuln_df)} vulnerabilities")
            
            if len(vuln_df) == 0:
                print("⚠ Warning: No vulnerabilities found in CSV")
                return False
            
            # Add severity labels for sorting
            vuln_df['Severity_Label'] = vuln_df['Severity'].apply(self.parse_severity)
            vuln_df['Severity_Order'] = vuln_df['Severity_Label'].map(self.SEVERITY_ORDER)
            
            # Sort by severity (Critical first, then High, Medium, Low)
            vuln_df = vuln_df.sort_values('Severity_Order')
            print(f"✓ Sorted by severity: Critical → High → Medium → Low")
            
            # Load template workbook
            wb = openpyxl.load_workbook(template_file)
            
            # Determine business unit
            if not business_unit and scan_info['asset_groups']:
                business_unit = scan_info['asset_groups']
            
            if not business_unit:
                business_unit = "Not Specified"
            
            # Update Details sheet
            self.update_details_sheet(wb, business_unit)
            
            # Get Vulnerability Details sheet
            if 'Vulnerability Details' not in wb.sheetnames:
                print("✗ ERROR: 'Vulnerability Details' sheet not found in template")
                return False
            
            ws = wb['Vulnerability Details']
            
            # Update business unit in header (legacy support)
            for row in ws.iter_rows(min_row=1, max_row=10):
                for cell in row:
                    if cell.value and 'Detailed Report for:' in str(cell.value):
                        cell.value = f"Detailed Report for: {business_unit}"
                        print(f"✓ Vulnerability Details: Business Unit header updated")
                        break
            
            # Find the header row
            header_row = 6
            
            # Add AutoFilter to headers for better usability
            try:
                ws.auto_filter.ref = f"B{header_row}:M{header_row}"
                print(f"✓ AutoFilter added to headers (Row {header_row})")
            except Exception as e:
                print(f"⚠ Warning: Could not add AutoFilter: {e}")
            
            # Clear existing data rows
            max_existing_row = ws.max_row
            if max_existing_row > header_row:
                ws.delete_rows(header_row + 1, max_existing_row - header_row)
            
            # Populate vulnerability data
            current_row = 7
            for idx, row in vuln_df.iterrows():
                # Column B: Asset IP Address
                ws.cell(row=current_row, column=2, value=self.clean_text(row.get('IP', '')))
                
                # Column C: Service Port
                port_protocol = self.format_port_protocol(row.get('Port', ''), row.get('Protocol', ''))
                ws.cell(row=current_row, column=3, value=port_protocol)
                
                # Column D: Vulnerability Severity Level
                severity = self.parse_severity(row.get('Severity', ''))
                ws.cell(row=current_row, column=4, value=severity)
                
                # Column E: Vulnerability CVSS Score
                cvss = self.get_cvss_score(row.get('Severity', ''))
                ws.cell(row=current_row, column=5, value=cvss)
                
                # Column F: Vulnerability CVE IDs
                cve_ids = self.clean_text(row.get('CVE ID', ''))
                ws.cell(row=current_row, column=6, value=cve_ids)
                
                # Column G: Vulnerability CVE URLs
                cve_urls = self.generate_cve_urls(row.get('CVE ID', ''))
                ws.cell(row=current_row, column=7, value=cve_urls)
                
                # Column H: Vulnerability Title
                ws.cell(row=current_row, column=8, value=self.clean_text(row.get('Title', '')))
                
                # Column I: Vulnerability Description (improved formatting)
                threat = self.clean_text(row.get('Threat', ''))
                impact = self.clean_text(row.get('Impact', ''))
                description = threat
                if impact and impact != threat:
                    description = f"{threat}\n\n{impact}" if threat else impact
                desc_cell = ws.cell(row=current_row, column=9, value=description)
                
                # Column J: Vulnerability Solution (improved formatting)
                solution = self.clean_text(row.get('Solution', ''))
                sol_cell = ws.cell(row=current_row, column=10, value=solution)
                
                # Column K: Vulnerability Proof
                proof_cell = ws.cell(row=current_row, column=11, value=self.clean_text(row.get('Results', '')))
                
                # Column L: Exploit Count
                exploit_count = self.count_exploits(row.get('Exploitability', ''), row.get('Bugtraq ID', ''))
                ws.cell(row=current_row, column=12, value=exploit_count)
                
                # Column M: Remediation Status
                ws.cell(row=current_row, column=13, value='Open')
                
                # Apply improved text wrapping and alignment
                for col in [9, 10, 11]:  # Description, Solution, Proof columns
                    cell = ws.cell(row=current_row, column=col)
                    cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
                
                # Set row height for better readability (auto-adjust based on content)
                ws.row_dimensions[current_row].height = 60  # Minimum height for readability
                
                current_row += 1
            
            print(f"✓ Populated {len(vuln_df)} vulnerabilities with improved formatting")
            
            # Create/Update Severity Tabulation sheet
            self.create_severity_tabulation_sheet(wb, vuln_df)
            
            # Create/Update Summary sheet
            self.update_summary_sheet(wb, vuln_df)
            
            # Reorder sheets to match desired sequence
            self.reorder_sheets(wb)
            
            # Save the output file
            wb.save(output_file)
            print(f"✓ Report saved to: {Path(output_file).name}")
            
            # Create monitoring log
            output_dir = Path(output_file).parent
            log_file = self.create_monitoring_log(csv_file, business_unit, vuln_df, output_dir)
            
            # Update statistics
            self.stats['successful'] += 1
            self.stats['total_vulnerabilities'] += len(vuln_df)
            
            return True
            
        except Exception as e:
            print(f"✗ ERROR processing file: {e}")
            import traceback
            traceback.print_exc()
            self.stats['failed'] += 1
            return False
    
    def create_severity_tabulation_sheet(self, wb, vuln_df):
        """Create and populate Severity Tabulation sheet with formulas"""
        
        # Check if sheet exists, if not create it
        if 'Severity Tabulation' not in wb.sheetnames:
            print("✓ Creating Severity Tabulation sheet...")
            ws = wb.create_sheet('Severity Tabulation')
        else:
            print("✓ Updating Severity Tabulation sheet...")
            ws = wb['Severity Tabulation']
        
        # Set column widths for better display
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
        
        # Add headers in Row 8
        headers = ['Asset Details', 'Critical', 'High', 'Medium', 'Low', 'Total']
        for col_idx, header in enumerate(headers, start=2):  # Start from column B (2)
            cell = ws.cell(row=8, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add Total row (Row 9) with formulas
        ws.cell(row=9, column=2, value='Total')
        ws.cell(row=9, column=2).font = Font(bold=True)
        
        # COUNTIF formulas for total row
        ws.cell(row=9, column=3, value='=COUNTIF(\'Vulnerability Details\'!$D:$D,"Critical")')
        ws.cell(row=9, column=4, value='=COUNTIF(\'Vulnerability Details\'!$D:$D,"High")')
        ws.cell(row=9, column=5, value='=COUNTIF(\'Vulnerability Details\'!$D:$D,"Medium")')
        ws.cell(row=9, column=6, value='=COUNTIF(\'Vulnerability Details\'!$D:$D,"Low")')
        ws.cell(row=9, column=7, value='=SUM(C9:F9)')
        
        # Extract unique IPs from vulnerability data
        unique_ips = sorted(vuln_df['IP'].unique())
        print(f"✓ Found {len(unique_ips)} unique IP addresses")
        
        # Populate IP list and formulas starting from Row 10
        for row_idx, ip in enumerate(unique_ips, start=10):
            # Column B: IP Address
            ws.cell(row=row_idx, column=2, value=ip)
            
            # Column C: Critical count for this IP
            ws.cell(row=row_idx, column=3, 
                   value=f'=COUNTIFS(\'Vulnerability Details\'!$B:$B,B{row_idx},\'Vulnerability Details\'!$D:$D,"Critical")')
            
            # Column D: High count for this IP
            ws.cell(row=row_idx, column=4,
                   value=f'=COUNTIFS(\'Vulnerability Details\'!$B:$B,B{row_idx},\'Vulnerability Details\'!$D:$D,"High")')
            
            # Column E: Medium count for this IP
            ws.cell(row=row_idx, column=5,
                   value=f'=COUNTIFS(\'Vulnerability Details\'!$B:$B,B{row_idx},\'Vulnerability Details\'!$D:$D,"Medium")')
            
            # Column F: Low count for this IP
            ws.cell(row=row_idx, column=6,
                   value=f'=COUNTIFS(\'Vulnerability Details\'!$B:$B,B{row_idx},\'Vulnerability Details\'!$D:$D,"Low")')
            
            # Column G: Total for this IP
            ws.cell(row=row_idx, column=7, value=f'=SUM(C{row_idx}:F{row_idx})')
        
        print(f"✓ Severity Tabulation sheet populated with {len(unique_ips)} IP addresses")
    
    def update_summary_sheet(self, wb, vuln_df):
        """Update Summary sheet - preserve existing chart, data updates automatically"""
        
        # Check if Summary sheet exists
        if 'Summary' not in wb.sheetnames:
            print("⚠ Warning: Summary sheet not found in template")
            print("  → Please add a Summary sheet with chart to your template")
            print("  → The chart should reference 'Severity Tabulation'!$C$8:$F$9")
            return
        
        ws = wb['Summary']
        has_existing_chart = len(ws._charts) > 0
        
        if has_existing_chart:
            print("✓ Summary sheet exists with chart - chart will auto-update from Severity Tabulation data")
        else:
            print("⚠ Warning: Summary sheet exists but no chart found")
            print("  → Please add a chart to the Summary sheet in your template")
    
    def reorder_sheets(self, wb):
        """Reorder sheets to match the desired sequence"""
        desired_order = ['Details', 'Vulnerability Details', 'Severity Tabulation', 'Summary', 'Glossary']
        
        # Get current sheets
        current_sheets = wb.sheetnames
        
        # Reorder based on desired sequence
        for position, sheet_name in enumerate(desired_order):
            if sheet_name in current_sheets:
                current_position = current_sheets.index(sheet_name)
                if current_position != position:
                    # Move sheet to correct position
                    wb.move_sheet(sheet_name, offset=position - current_position)
                    # Update current_sheets list
                    current_sheets = wb.sheetnames
        
        print(f"✓ Sheets reordered: {' → '.join(wb.sheetnames)}")
    
    def create_monitoring_log(self, csv_file, business_unit, vuln_df, output_dir):
        """Create a monitoring/logging Excel file with scan summary"""
        
        try:
            # Extract scan information
            scan_info = self.extract_scan_info(csv_file)
            
            # Parse scan metadata
            scan_title = scan_info.get('scan_title', '')
            scan_date = scan_info.get('scan_date', '')
            
            # Extract CRITICALITY from first word of scan title (e.g., "INTERNAL_CCaaS..." -> "INTERNAL")
            criticality = scan_title.split('_')[0].split()[0].upper() if scan_title else 'UNKNOWN'
            
            # Parse frequency from title (look for Q1, Q2, Q3, Q4, or assume Annually)
            frequency = 'Quarterly' if any(q in scan_title.upper() for q in ['Q1', 'Q2', 'Q3', 'Q4']) else 'Annually'
            
            # Extract asset statistics from CSV header
            active_hosts = 0
            total_hosts = 0
            
            try:
                with open(csv_file, 'r', encoding='utf-8') as f:
                    lines = [f.readline() for _ in range(7)]
                    if len(lines) > 5:
                        parts = lines[5].strip('"').split('","')
                        if len(parts) >= 3:
                            active_hosts = int(parts[1]) if parts[1].isdigit() else 0
                            total_hosts = int(parts[2]) if parts[2].isdigit() else 0
            except Exception as e:
                print(f"⚠ Warning: Could not extract asset counts: {e}")
            
            # Count vulnerabilities by severity
            severity_counts = {
                'Critical': 0,
                'High': 0,
                'Medium': 0,
                'Low': 0
            }
            
            for _, row in vuln_df.iterrows():
                severity = self.parse_severity(row.get('Severity', ''))
                if severity in severity_counts:
                    severity_counts[severity] += 1
            
            # Count unique IPs with findings
            assets_with_findings = len(vuln_df['IP'].unique()) if len(vuln_df) > 0 else 0
            
            # Calculate unreachable assets
            assets_not_scanned = total_hosts - active_hosts
            
            # Prepare log data
            log_data = {
                'BU NAME': business_unit,
                'CRITICALITY': criticality,
                'FREQUENCY': frequency,
                'REPORT ISSUE': datetime.strptime(scan_date, '%m/%d/%Y at %H:%M:%S (GMT%z)').strftime('%d-%b-%y') if scan_date else datetime.now().strftime('%d-%b-%y'),
                'ASSIGNED': 'ePLDT CSOG - Risk, Compliance & Vulnerability',
                'TOTAL # OF ASSETS IN SCOPE (AFTER SCAN)': total_hosts,
                'TOTAL # OF ASSET SCANNED/REACHABLE': active_hosts,
                'TOTAL # OF ASSETS NOT SCANNED (UNREACHABLE)': assets_not_scanned,
                'TOTAL # OF ASSETS WITH FINDINGS (SCANNING)': assets_with_findings,
                'DATE OF INITIAL RELEASE': datetime.now().strftime('%d-%b-%y'),
                'CRITICAL': severity_counts['Critical'],
                'HIGH': severity_counts['High'],
                'MEDIUM': severity_counts['Medium'],
                'LOW': severity_counts['Low'],
                'TOTAL': sum(severity_counts.values())
            }
            
            # Create monitoring log workbook
            log_wb = openpyxl.Workbook()
            log_ws = log_wb.active
            log_ws.title = "Scan Monitoring Log"
            
            # Write headers (Row 1)
            headers = list(log_data.keys())
            for col_idx, header in enumerate(headers, start=1):
                cell = log_ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Apply special formatting to "INITIAL VA RESULTS" columns (Critical, High, Medium, Low)
            for col_idx in range(11, 15):  # Columns K, L, M, N (Critical, High, Medium, Low)
                cell = log_ws.cell(row=1, column=col_idx)
                cell.fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
            
            # Write data (Row 2)
            for col_idx, header in enumerate(headers, start=1):
                log_ws.cell(row=2, column=col_idx, value=log_data[header])
            
            # Set column widths
            column_widths = {
                'A': 45,  # BU NAME
                'B': 12,  # CRITICALITY
                'C': 12,  # FREQUENCY
                'D': 15,  # REPORT ISSUE
                'E': 45,  # ASSIGNED
                'F': 20,  # TOTAL # OF ASSETS IN SCOPE
                'G': 20,  # TOTAL # OF ASSET SCANNED
                'H': 20,  # TOTAL # OF ASSETS NOT SCANNED
                'I': 25,  # TOTAL # OF ASSETS WITH FINDINGS
                'J': 18,  # DATE OF INITIAL RELEASE
                'K': 10,  # CRITICAL
                'L': 10,  # HIGH
                'M': 10,  # MEDIUM
                'N': 10,  # LOW
                'O': 10   # TOTAL
            }
            
            for col_letter, width in column_widths.items():
                log_ws.column_dimensions[col_letter].width = width
            
            # Set row height for header
            log_ws.row_dimensions[1].height = 30
            
            # Generate log filename
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            csv_basename = Path(csv_file).stem
            log_filename = f"Monitoring_Log_{csv_basename}_{timestamp}.xlsx"
            log_filepath = Path(output_dir) / log_filename
            
            # Save the log file
            log_wb.save(log_filepath)
            print(f"✓ Monitoring log created: {log_filename}")
            
            return str(log_filepath)
            
        except Exception as e:
            print(f"⚠ Warning: Could not create monitoring log: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def batch_process(self, csv_dir, template_file, output_dir, business_unit=None):
        """Process multiple CSV files in a directory"""
        csv_files = list(Path(csv_dir).glob('*.csv'))
        
        if not csv_files:
            print(f"No CSV files found in {csv_dir}")
            return
        
        print(f"\n{'='*60}")
        print(f"BATCH PROCESSING MODE")
        print(f"{'='*60}")
        print(f"Found {len(csv_files)} CSV file(s)")
        print(f"Template: {Path(template_file).name}")
        print(f"Output directory: {output_dir}")
        
        # Create output directory if it doesn't exist
        Path(output_dir).mkdir(parents=True, exist_ok=True)
        
        self.stats['total_files'] = len(csv_files)
        
        for csv_file in csv_files:
            # Generate output filename
            csv_name = csv_file.stem
            output_name = f"Detailed_Report_{csv_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            output_file = Path(output_dir) / output_name
            
            self.fill_vulnerability_details(
                template_file,
                str(csv_file),
                str(output_file),
                business_unit
            )
        
        # Print summary
        self.print_summary()
    
    def print_summary(self):
        """Print processing summary"""
        print(f"\n{'='*60}")
        print(f"PROCESSING SUMMARY")
        print(f"{'='*60}")
        print(f"Total files processed: {self.stats['total_files']}")
        print(f"Successful: {self.stats['successful']}")
        print(f"Failed: {self.stats['failed']}")
        print(f"Total vulnerabilities: {self.stats['total_vulnerabilities']}")
        print(f"{'='*60}\n")

def main():
    automation = QualysReportAutomation()
    
    if len(sys.argv) < 2:
        print("""
╔══════════════════════════════════════════════════════════════╗
║   Qualys Vulnerability Report Automation Tool v3.2          ║
╚══════════════════════════════════════════════════════════════╝

USAGE:
  Single File Mode:
    python qualys_report_automation_v3.py <csv_file> <template> <output> [business_unit]
  
  Batch Processing Mode:
    python qualys_report_automation_v3.py --batch <csv_dir> <template> <output_dir> [business_unit]

EXAMPLES:
  Single:
    python qualys_report_automation_v3.py scan.csv template.xlsx output.xlsx "ePLDT, Inc."
  
  Batch:
    python qualys_report_automation_v3.py --batch ./scans/ template.xlsx ./reports/ "ePLDT, Inc."

ARGUMENTS:
  csv_file      - Path to Qualys CSV scan file
  csv_dir       - Directory containing multiple CSV files
  template      - Path to Excel template file
  output        - Path for output Excel file (or directory for batch)
  business_unit - Business Unit name (auto-detected from CSV if not provided)

NEW IN v3.2:
  ✓ Monitoring/logging Excel file automatically generated
  ✓ Extracts scan metadata for tracking and reporting
  ✓ Copy-paste ready format for monitoring dashboards

v3.1 FEATURES:
  ✓ Summary sheet & chart preserved from template (no recreation)
  ✓ Chart auto-updates from Severity Tabulation data
  ✓ Simplified workflow - just update data, chart refreshes

v3.0 FEATURES:
  ✓ Fixed Business Unit field update in Details sheet
  ✓ Added AutoFilter to Vulnerability Details headers
  ✓ Improved text formatting for better readability
  ✓ Enhanced row height management
  ✓ Better error handling and validation
        """)
        sys.exit(1)
    
    # Check for batch mode
    if sys.argv[1] == '--batch':
        if len(sys.argv) < 5:
            print("ERROR: Batch mode requires: --batch <csv_dir> <template> <output_dir> [business_unit]")
            sys.exit(1)
        
        csv_dir = sys.argv[2]
        template_file = sys.argv[3]
        output_dir = sys.argv[4]
        business_unit = sys.argv[5] if len(sys.argv) > 5 else None
        
        automation.batch_process(csv_dir, template_file, output_dir, business_unit)
    else:
        # Single file mode
        if len(sys.argv) < 4:
            print("ERROR: Single mode requires: <csv_file> <template> <output> [business_unit]")
            sys.exit(1)
        
        csv_file = sys.argv[1]
        template_file = sys.argv[2]
        output_file = sys.argv[3]
        business_unit = sys.argv[4] if len(sys.argv) > 4 else None
        
        automation.stats['total_files'] = 1
        success = automation.fill_vulnerability_details(template_file, csv_file, output_file, business_unit)
        automation.print_summary()
        
        sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()
